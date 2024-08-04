import openpyxl
import xlwt
import os
import csv
import json

def set_style(name,size,color,borders_color, top, color_fore,blod=False):
    style = xlwt.XFStyle()  # 初始化样式
    # 字体
    font = xlwt.Font()
    font.name = name
    font.height = 20 * size  # 字号
    font.bold = blod  # 加粗
    font.colour_index = color  # 默认：0x7FFF 黑色：0x08
    style.font = font
    # 居中
    alignment = xlwt.Alignment()  # 居中
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    alignment.vert = xlwt.Alignment.VERT_CENTER
    style.alignment=alignment
    # 边框
    borders = xlwt.Borders()
    if top:
        borders.top = xlwt.Borders.THIN
        borders.top_colour = borders_color
    borders.bottom = xlwt.Borders.THIN
    borders.bottom_colour = borders_color
    borders.left = xlwt.Borders.THIN
    # borders.b_colour = borders_color
    borders.right = xlwt.Borders.THIN
    # borders.bottom_colour = borders_color
    style.borders = borders
    # 背景颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # 设置背景颜色的模式(NO_PATTERN; SOLID_PATTERN)
    pattern.pattern_fore_colour = color_fore  # 默认：无色：0x7FFF；黄色：0x0D；蓝色：0x0C
    style.pattern = pattern
    return style

def city_name_equa(a:str, b:str, f = 1):
    flag = 1
    if len(a) > 2:
        a = a[:-1]
    if len(b) > 2:
        b = b[:-1]
    for i in a:
        if i not in b:
            flag = 0
            break
    if flag or (f and city_name_equa(b, a, 0)):
        return 1 # 双向属于满足其一即匹配成功
    else:
        return 0 # 匹配失败

path = os.path.split(os.path.realpath(__file__))[0]
city_name_lut = {}
with open(path[:-10] + "ChinaCityList.json", encoding='utf-8') as loadf:
    ori_city_num_lut = json.load(loadf)
    for province in ori_city_num_lut:
        for city in province['cities']:
            city_name_lut[city['code']] = city['name']
# print(city_name_lut)

name_wnum_lut = openpyxl.load_workbook(path[:-10] + "气象站编号对照表.xlsx") 
name_wnum_lut = name_wnum_lut.active
name_city = []
province_name = []
weather_number = []
city_number = []
for row in name_wnum_lut.iter_rows(min_row=2, min_col=1, max_row=name_wnum_lut.max_row, max_col=5, values_only=True):
    try:
        with open(path[:-10] + "2023ori_data\\" + str(row[0]) + "099999.csv", 'r') as file:
            ori_data = csv.reader(file)
            pass
    except:
        continue
    if row[3] != None:
        this_name = row[3]
    elif row[4] != None:
        this_name = row[4]
    else:
        this_name = row[1]
    if this_name not in name_city:
        name_city.append(this_name)
        province_name.append(row[2])
        weather_number.append(row[0])
        for key in city_name_lut:
            if city_name_equa(city_name_lut[key], this_name):
                city_number.append(key)
                name_city[-1] = city_name_lut[key]
                break
        else:
            city_number.append('error')
# print(len(name_city))
# print(name_city)
# print(city_number)

esc30 = [] # 最高气温86℉以上的累计天数
esc35 = [] # 最高气温95℉以上的累计天数
esc40 = [] # 最高气温104℉以上的累计天数
bel10 = [] # 最低气温14华氏度以下的累计天数
bel20 = [] # 最低气温-4华氏度以下的累计天数
amtd = [] # 年均昼夜温差 ℃
amws = [] # 年均风速 m/s
atp = [] # 年总降水量 mm
mem = [] # 平均温度最适区间[50℉,68℉]隶属度和 单日隶属度=1,50<x<68;-0.003324x^2+0.3922x-10.3711,other 求年和 x为日平均气温
for i in range(len(weather_number)):
    esc30.append(0)
    esc35.append(0)
    esc40.append(0)
    bel10.append(0)
    bel20.append(0)
    amtd.append(0)
    amws.append(0)
    atp.append(0)
    mem.append(0)
    for year in range(2019, 2024):
        with open(path[:-10] + str(year) + "ori_data\\" + str(weather_number[i]) + "099999.csv", 'r') as file:
            ori_data = csv.reader(file)
            next(ori_data)
            for day in ori_data:
                if float(day[20]) > 86:
                    esc30[i] += 1
                if float(day[20]) > 95:
                    esc35[i] += 1
                if float(day[20]) > 104:
                    esc40[i] += 1
                if float(day[22]) < 14:
                    bel10[i] += 1
                if float(day[22]) < -4:
                    bel20[i] += 1
                amtd[i] += float(day[20]) - float(day[22])
                amws[i] += 0.514 * float(day[16])
                atp[i] += (25.4 * float(day[24])) if day[24] != '99.99' else 0
                if float(day[6]) < 68 and float(day[6]) > 50:
                    mem[i] += 1
                else:
                    mem[i] += max(-0.003324 * (float(day[6])**2) + 0.3922 * float(day[6]) - 10.3711, 0)
    amtd[i] = (amtd[i] - 32) / 1.8 / 365 / 5
    amws[i] /= 365 * 5
    mem[i] /= 5
    esc30[i] /= 5
    esc35[i] /= 5
    esc40[i] /= 5
    bel10[i] /= 5
    bel20[i] /= 5
    atp[i] /= 5

points = []
data = [name_city, province_name, city_number, weather_number, amtd, esc30, esc35, esc40, bel10, bel20, amws, atp, mem, points]
# 超30 超35 超40 低10 低20 风速 降水量 舒适天数
sub_weight = (0.173, 0.354, 0.173, 0.032, 0.035, 0.028, 0.028, 0.173)
obj_weight = (0.167, 0.067, 0.037, 0.209, 0.081, 0.015, 0.170, 0.249)

for p in data[5:-1]:
    p.append(max(p))
    p.append(min(p))

for i in range(len(data[0])):
    points.append(0)
    for j in range(7):
        points[-1] += (0.8 * sub_weight[j] + 0.2 * obj_weight[j]) * ((data[j+5][-1] - data[j+5][i]) / (data[j+5][-2] - data[j+5][-1]) + 1) * 100 // 1
    points[-1] += (0.8 * sub_weight[7] + 0.2 * obj_weight[7]) * (data[12][i] - data[12][-1]) / (data[12][-2] - data[12][-1]) * 100 // 1


res = xlwt.Workbook(encoding = 'utf-8')
sheet = res.add_sheet('sheet1')
sheet.row(0).height_mismatch = True
sheet.row(0).height = 500
title = ['城市', '省', '行政区代码', '气象站代码', '均温差', '超30', '超35', '超40', '低10', '低20', '风速', '降水量', '舒适天数', '综合评分']

for i in range(len(title)):
    sheet.write(0, i, title[i], set_style('黑体', 11, 0x7FFF, 0x00, 1, 0x7FFF, blod = True))
    if i < 2:
        sheet.col(i).width = 6000
    elif i < 4:
        sheet.col(i).width = 4000
    else:
        sheet.col(i).width = 3000
style_cont = set_style('宋体', 10, 0x7FFF, 0x16, 0, 0x7FFF, blod = False)
style_separate = set_style('宋体', 10, 0x7FFF, 0x00, 0, 0x7FFF, blod = False)
for i in range(len(name_city)):
    for row in range(len(data)):
        sheet.row(i+1).height_mismatch = True
        sheet.row(i+1).height = 400
        sheet.write(i+1, row, data[row][i] if type(data[row][i]) != float else int(data[row][i]), style_cont)
res.save('气象数据.xls')
# print(name_city)
# print(esc30)

